from openpyxl import load_workbook
from io import BytesIO


def extract_text_from_xlsx(content: bytes) -> str:
    """
    Extracts text from an XLSX file.

    Args:
        content: The binary content of the XLSX file

    Returns:
        Extracted text as a string
    """
    text_total = ""

    # Load the workbook from binary content
    in_memory_file = BytesIO(content)
    workbook = load_workbook(in_memory_file, data_only=True)

    # Process each sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join([str(cell) for cell in row if cell is not None])
            text_total += row_text + "\n"

    return text_total
